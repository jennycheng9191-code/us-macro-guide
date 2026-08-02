"""NY Fed 消費者預期調查（SCE）。

結構已於 2026-08-02 實地確認：
工作表 'Inflation expectations'，第 3 列為表頭、第 4 列起為資料，
第 0 欄為 YYYYMM、第 1 欄為 1 年期中位數、第 2 欄為 3 年期中位數。
卡片主值採用 3 年期（Fed 引用度最高，短端易受汽油價干擾）。
"""
from __future__ import annotations

import io

import openpyxl
import requests

from common import UA


def fetch(card_id: str, m: dict) -> dict:
    r = requests.get(m["url"], timeout=60, headers={"User-Agent": UA})
    r.raise_for_status()
    wb = openpyxl.load_workbook(io.BytesIO(r.content), read_only=True, data_only=True)
    if "Inflation expectations" not in wb.sheetnames:
        return {"ok": False, "reason": "SCE 檔案缺少 'Inflation expectations' 工作表"}
    ws = wb["Inflation expectations"]

    hist, one_year = [], None
    for row in ws.iter_rows(min_row=5, values_only=True):
        if not row or row[0] is None:
            continue
        ym = str(row[0]).strip()
        if len(ym) != 6 or not ym.isdigit():
            continue
        try:
            three = float(row[2])
            one_year = float(row[1])
        except (TypeError, ValueError, IndexError):
            continue
        hist.append({"date": f"{ym[:4]}-{ym[4:]}-01", "value": three})

    if not hist:
        return {"ok": False, "reason": "SCE 工作表未解析出任何資料列"}

    return {"ok": True, "value": hist[-1]["value"], "asof": hist[-1]["date"],
            "history": hist[-24:], "raw_latest": hist[-1]["value"], "freq": "M",
            "extras": {"1年期中位數": one_year},
            "also": {}, "source_label": "NY Fed SCE 官方資料檔",
            "value_label": "3年期中位數"}
